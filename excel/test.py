from openpyxl import load_workbook, Workbook
from openpyxl.compat import range

data = load_workbook('file_test.xlsx')
print(data.sheetnames)
print('*'*50)
sheet = data['s_1']

#считываем данные с ячейки
val = sheet['D2'].value
#print(val)

# диапазон ячеек
#B = [i[0].value for i in sheet.range('A1:D2')]
A = [i[j].value for i in sheet['A1':'D2'] for j in range(4)]
print(A)
print('*'*50)
for i in sheet['A1':'D2']:
    for j in i:
        print(j.value, end=' ')
    print()

print('*'*50)
C=[]
for i in range(1,8):
    C.append(sheet.cell(row=i,column=2).value)
print(C)
#записываем данные в ячейку
sheet['E1']= 7
sheet['E2']= 'e'
data.save('file_test.xlsx')

#записываем данные в диапазон ячеек
for i in range(3,8):
    sheet.cell(row=i, column=2).value = i*i
data.save('file_test.xlsx')
