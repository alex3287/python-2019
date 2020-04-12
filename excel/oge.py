# Данная прога обрабатывает excel файлы с расширением xlsx
# + и - будут переведены в 1 и 0 и расставлены отдельно по клеточкам
# программа была написана на python3.6, для корректной работы
# потребуется установленная библиотека openpyxl

import os
import openpyxl

# если файл result_oge.xlsx отсутствует, то создаем его
f = os.listdir()
if 'result_oge.xlsx' not in f:
    book = openpyxl.Workbook()
    book.save('result_oge.xlsx')

# считываем содержимое папки где находятся все файлы
Files = ['./oge_files/'+i for i in os.listdir('./oge_files') if 'xlsx' in i]
#print(Files)

def open_xlsx(file):
    '''open file excel and read name OU'''
    file = openpyxl.load_workbook(file)
    sheet = file[file.sheetnames[0]]
    name = sheet['A4'].value
    n = name.find('№')
    nameOU = name[n+1:].strip()
    if len(nameOU) > 10:
        nameOU = nameOU[:2]
    if not nameOU[-1].isdigit():
        nameOU = nameOU[:-1]
    return nameOU

def marks(file):
    '''open file excel and read marks column 15 row 8'''
    file2 = openpyxl.load_workbook(file)
    sheet = file2[file2.sheetnames[0]]
    i=8
    Data=[]
    while sheet.cell(row=i,column=13).value:
        Data.append(sheet.cell(row=i,column=15).value)
        i+=1
    two = three = four = five =  0
    two = Data.count(2)
    three = Data.count(3)
    four = Data.count(4)
    five = Data.count(5)
    school = open_xlsx(file)
    return school, two, three, four, five

def get_last_name(file):
    '''open file excel and read last name column 7 row 8'''
    file = openpyxl.load_workbook(file)
    sheet = file[file.sheetnames[0]]
    i=8
    Data=[]
    while sheet.cell(row=i,column=7).value:
        Data.append(sheet.cell(row=i,column=7).value)
        i+=1
    return Data

def read_answer(file):
    '''open file excel and read answer column 12 row 8'''
    file = openpyxl.load_workbook(file)
    sheet = file[file.sheetnames[0]]
    i=8
    Data=[]
    while sheet.cell(row=i,column=12).value:
        Data.append(sheet.cell(row=i,column=12).value)
        i+=1
    return Data

def read_answer2(file):
    '''open file excel and read answer 2 column 13 row 8'''
    file = openpyxl.load_workbook(file)
    sheet = file[file.sheetnames[0]]
    i=8
    Data=[]
    while sheet.cell(row=i,column=12).value:
        s = sheet.cell(row=i, column=13).value
        s2 = s[0]+s[4]
        Data.append(s2)
        i+=1
    return Data

def add_result(file):
    '''open file and add new sheet with zeros and ones (answer 1)'''
    book = openpyxl.load_workbook(file)
    sheet = book.create_sheet(title='result')
    answer = read_answer(file)
    count = len(answer)
    for k in range(1,19):
        sheet.cell(row=1, column=k+1).value = k
    for k2 in range(2,count+2):
        sheet.cell(row=k2, column=1).value = k2-1
    for i in range(count):
        data = answer[i]
        for j in range(18):
            if data[j] == '+':
                sheet.cell(row=i+2, column=j+2).value = 1
            else:
                sheet.cell(row=i + 2, column=j + 2).value = 0
    book.save(file)

def add_result2(file):
    '''open file and add new sheet with answer 2'''
    book = openpyxl.load_workbook(file)
    sheet = book.create_sheet(title='result_2')
    answer = read_answer2(file)
    count = len(answer)
    for k in range(1,3):
        sheet.cell(row=1, column=k+1).value = k
    for j in range(2,count+2):
        sheet.cell(row=j, column=1).value = j-1
    for i in range(count):
        data = answer[i]
        sheet.cell(row=i+2, column=2).value = int(data[0])
        sheet.cell(row=i + 2, column=3).value = int(data[1])
    book.save(file)

def collect_data(file):
    '''open file and take data, add new sheet
     and fill zeros and ones (answer 1)'''
    book = openpyxl.load_workbook('result_oge.xlsx')
    name = open_xlsx(file)
    sheet = book.create_sheet(title=name)
    answer = read_answer(file)
    count = len(answer)
    last_names = get_last_name(file)
    for k in range(1,19):
        sheet.cell(row=1, column=k+2).value = k
    for k2 in range(2,count+2):
        sheet.cell(row=k2, column=1).value = k2-1
    sheet.cell(row=1, column=2).value = 'Фамилия'
    for n in range(count):
        sheet.cell(row=n + 2, column=2).value = last_names[n]
    for i in range(count):
        data = answer[i]
        for j in range(18):
            if data[j] == '+':
                sheet.cell(row=i+2, column=j+3).value = 1
            else:
                sheet.cell(row=i + 2, column=j+3).value = 0
    book.save('result_oge.xlsx')

def collect_data_2(file):
    '''open file and add new sheet with answer 2'''
    book = openpyxl.load_workbook('result_oge.xlsx')
    name = open_xlsx(file)
    name += '_2'
    sheet = book.create_sheet(title=name)
    answer = read_answer2(file)
    last_names = get_last_name(file)
    count = len(answer)
    for k in range(1,3):
        sheet.cell(row=1, column=k+2).value = k
    for j in range(2,count+2):
        sheet.cell(row=j, column=1).value = j-1
    sheet.cell(row=1, column=2).value = 'Фамилия'
    for n in range(count):
        sheet.cell(row=n + 2, column=2).value = last_names[n]
    for i in range(count):
        data = answer[i]
        sheet.cell(row=i+2, column=3).value = int(data[0])
        sheet.cell(row=i + 2, column=4).value = int(data[1])
    book.save('result_oge.xlsx')

def all_collect(Files):
    '''open all files and take data, add new file
     and fill zeros and ones (answer 1)'''
    book = openpyxl.load_workbook('result_oge.xlsx')
    sheet = book.create_sheet(title='all_answer_1')
    answers = []
    for file in Files:
        answers += read_answer(file)
    count = len(answers)
    for k in range(1,19):
        sheet.cell(row=1, column=k+1).value = k
    for k2 in range(2,count+2):
        sheet.cell(row=k2, column=1).value = k2-1
    for i in range(count):
        data = answers[i]
        for j in range(18):
            if data[j] == '+':
                sheet.cell(row=i+2, column=j+2).value = 1
            else:
                sheet.cell(row=i + 2, column=j + 2).value = 0
    book.save('result_oge.xlsx')

def all_collect_2(Files):
    '''open file and add new sheet with answer 2'''
    book = openpyxl.load_workbook('result_oge.xlsx')
    sheet = book.create_sheet(title='all_answer_2')
    answers = []
    for file in Files:
        answers += read_answer2(file)
    count = len(answers)
    for k in range(1,3):
        sheet.cell(row=1, column=k+1).value = k
    for j in range(2,count+2):
        sheet.cell(row=j, column=1).value = j-1
    for i in range(count):
        data = answers[i]
        sheet.cell(row=i+2, column=2).value = int(data[0])
        sheet.cell(row=i + 2, column=3).value = int(data[1])
    book.save('result_oge.xlsx')

if __name__=='__main__':
    print('''
            Данная прога обрабатывает excel файлы с расширением xlsx
             + и - будут переведены в 1 и 0 и расставлены отдельно по клеточкам.
             Также программа может выдавать статистику по оценками
    (программа была написана на python3.6, для корректной работы потребуется установленная библиотека openpyxl)
    
                Для продолжнения работы выбирите нужное вам действие
                    1 - будут сформированы новые листы в каждом файле с результатами
                    2 - будут сформированы данные в отдельные листы одного общего файла (листы будут названы по школам)
                    3 - данные будут собраны на двух листах отдельного файла (ответы 1 части и 2)
                    4 - выведит на экран общее количество оценок
                    5 - выведит на экран оценки по школам отдельно
                    6 - выведит на экран количество сдающих по школам отдельно и средний балл
                    7 - общее количество сдающих, средний балл
                    8 - test
                    для выхода из программы нажмите любу клавишу отличную от предыдущих
                    пока над этим думаю))
                   
    ''')
    choice = input('Сделайте свой выбор -> ')
    if choice == '1':
        for i in Files:
            add_result(i)
            add_result2(i)
        print('successful')
    elif choice == '2':
        time = 0
        for i in Files:
            time+=1
            print(time*5,'%')
            collect_data(i)
            time += 1
            print(time * 5, '%')
            collect_data_2(i)
        print('successful')
    elif choice == '3':
        all_collect(Files)
        all_collect_2(Files)
        print('successful')
    elif choice == '4':
        t = th = f = fi = 0
        for i in Files:
            s, t1, th1, f1, fi1 = marks(i)
            t += t1
            th += th1
            f += f1
            fi += fi1
        print('Общее число оценок по всем школам\n',2,'-',t,'\n',3,'-',th,'\n',4,'-',f,'\n',5,'-',fi)
    elif choice == '5':
        for i in Files:
            s, t, th, f, fi = marks(i)
            print('Школа №', s, '\n', 2, '-', t, '\n', 3, '-', th, '\n', 4, '-', f, '\n', 5, '-', fi)
    elif choice == '6':
        for i in Files:
            s, t, th, f, fi = marks(i)
            pupil = t+th+f+fi
            mark = t*2+th*3+f*4+fi*5
            print('Школа №', s, '-', pupil,'\nсредний балл',mark/pupil)
    elif choice == '7':
        students = 0
        sum_marks = 0
        for i in Files:
            s, t, th, f, fi = marks(i)
            students += t+th+f+fi
            sum_marks += t*2 + th*3 + f*4 + fi*5
        average = sum_marks / students
        print('общее количество учеников -',students)
        print('средний балл -', average)
        print(sum_marks)
    elif choice == '8':
        print(get_last_name(Files[0]))
    else:
        print('Good by)')